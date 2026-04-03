#!/usr/bin/env python3
"""
跨境工具 A/B 测试 Benchmark

Phase 1: 固定搜索=百度千帆，对比 Sonnet 4.6 vs Haiku 4.5
Phase 2: 固定 LLM，对比 7 种搜索 API 组合（控制总检索量≈12条）

用法:
  /d/Space/Anaconda3_Space/envs/cb_tool/python.exe experiments/benchmark.py --phase 1
  /d/Space/Anaconda3_Space/envs/cb_tool/python.exe experiments/benchmark.py --phase 2 --llm haiku
"""
import argparse
import asyncio
import logging
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from backend.app.core.config import settings
from backend.app.services.search.baidu_qianfan import BaiduQianfanProvider
from backend.app.services.search.volcengine import VolcengineProvider
from backend.app.services.search.metaso import MetasoProvider
from backend.app.services.llm.anthropic_compat import AnthropicCompatProvider
from backend.app.services.extract.extractor import extract_company_info
from backend.app.services.extract.fields import SUMMARY_FIELDS
from backend.app.services.pipeline.keywords import build_keyword_groups
from backend.app.services.pipeline.runner import _multi_search

logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger("benchmark")
logger.setLevel(logging.INFO)

TEST_COMPANIES = [
    "深圳市有芯电子有限公司",
    "深圳市鹏润米业有限公司",
    "深圳市众阳电路科技有限公司",
    "佛山市元通胶粘实业有限公司",
    "深圳市量子数据科技有限公司",
    "广州海纳数智技术有限公司",
    "深圳市思桅电子有限公司",
    "广东游盟科技股份有限公司",
    "深圳市多乐声电子有限公司",
    "上海丽邱缘科技有限公司",
    "广东中新实业有限公司",
    "苏州托玛斯机器人集团有限公司",
    "深圳市德奥信息技术有限公司",
    "北京云杉世界信息技术集团有限公司",
    "江苏容轩光电科技有限公司",
    "深圳易速马网络科技有限公司",
]

# ── LLM 模型配置（如代理不支持这些 model ID，请手动修改）──────────────────
LLM_CONFIGS = {
    "sonnet": {"model": "claude-sonnet-4-6", "label": "Sonnet4.6"},
    "haiku":  {"model": "claude-haiku-4-5",  "label": "Haiku4.5"},
}

# 搜索条数控制：单源12条，双源各6条，三源各4条 → 总量≈12
SEARCH_LIMITS = {1: 12, 2: 6, 3: 4}


# ── 计数 Wrapper ──────────────────────────────────────────────────────────

class CountingSearch:
    """包装 SearchProvider，统计调用次数"""
    def __init__(self, provider):
        self._p = provider
        self.call_count = 0

    @property
    def name(self):
        return self._p.name

    async def search(self, query, num_results=10):
        self.call_count += 1
        return await self._p.search(query, num_results)

    async def close(self):
        await self._p.close()


class CountingLLM:
    """包装 LLMProvider，累计 token 消耗"""
    def __init__(self, provider, label):
        self._p = provider
        self.label = label
        self.total_input = 0
        self.total_output = 0
        self.call_count = 0

    async def complete(self, system_prompt, user_prompt, **kw):
        kw.setdefault("max_tokens", 4096)
        resp = await self._p.complete(system_prompt, user_prompt, **kw)
        self.call_count += 1
        self.total_input += resp.prompt_tokens
        self.total_output += resp.completion_tokens
        return resp

    async def close(self):
        await self._p.close()


# ── 单次组合运行 ──────────────────────────────────────────────────────────

@dataclass
class RunResult:
    label: str
    total_time_s: float
    search_calls: dict          # {provider_name: count}
    llm_input_tokens: int
    llm_output_tokens: int
    llm_calls: int
    rows: list                  # [{公司名称, 字段值, 备注...}]


async def run_combination(
    label: str,
    search_with_limits: list,   # [(CountingSearch, num_results)]
    llm: CountingLLM,
) -> RunResult:
    rows = []
    t0 = time.monotonic()

    for i, company in enumerate(TEST_COMPANIES, 1):
        logger.info("[%s] (%d/%d) %s", label, i, len(TEST_COMPANIES), company)

        # 搜索阶段
        all_results = []
        for query in build_keyword_groups(company):
            merged, _ = await _multi_search(query, search_with_limits)
            all_results.extend(merged)

        # LLM 抽取阶段
        extraction, _, _ = await extract_company_info(
            company, all_results, llm, SUMMARY_FIELDS
        )

        row = {"公司名称": company, "证据条数": extraction.evidence_count}
        for fd in SUMMARY_FIELDS:
            row[fd.col_name] = extraction.summary.get(fd.id, "")
            row[fd.col_name + "_备注"] = extraction.summary_notes.get(fd.id, "")
        if extraction.error:
            row["错误"] = extraction.error
        rows.append(row)

    elapsed = time.monotonic() - t0
    logger.info("[%s] 完成，耗时 %.1fs", label, elapsed)

    return RunResult(
        label=label,
        total_time_s=elapsed,
        search_calls={p.name: p.call_count for p, _ in search_with_limits},
        llm_input_tokens=llm.total_input,
        llm_output_tokens=llm.total_output,
        llm_calls=llm.call_count,
        rows=rows,
    )


# ── Excel 输出 ────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
CELL_FONT = Font(size=10)


def _build_columns():
    cols = ["公司名称"]
    for fd in SUMMARY_FIELDS:
        cols.append(fd.col_name)
        cols.append(fd.col_name + "_备注")
    cols.extend(["证据条数", "错误"])
    return cols


def _write_result_sheet(ws, result: RunResult):
    cols = _build_columns()
    for c, name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for r, row_data in enumerate(result.rows, 2):
        for c, col_name in enumerate(cols, 1):
            cell = ws.cell(row=r, column=c, value=row_data.get(col_name, ""))
            cell.font = CELL_FONT

    for c, col_name in enumerate(cols, 1):
        letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[letter].width = max(len(col_name) * 2, 12)


def _write_metrics_sheet(ws, results: list):
    all_search_names = []
    for r in results:
        for name in r.search_calls:
            if name not in all_search_names:
                all_search_names.append(name)

    headers = ["组合", "总耗时(s)", "LLM调用次数",
               "Input Tokens", "Output Tokens"]
    for name in all_search_names:
        headers.append(f"搜索调用({name})")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for r_idx, result in enumerate(results, 2):
        ws.cell(row=r_idx, column=1, value=result.label).font = CELL_FONT
        ws.cell(row=r_idx, column=2, value=round(result.total_time_s, 1)).font = CELL_FONT
        ws.cell(row=r_idx, column=3, value=result.llm_calls).font = CELL_FONT
        ws.cell(row=r_idx, column=4, value=result.llm_input_tokens).font = CELL_FONT
        ws.cell(row=r_idx, column=5, value=result.llm_output_tokens).font = CELL_FONT
        for i, name in enumerate(all_search_names):
            ws.cell(row=r_idx, column=6 + i,
                    value=result.search_calls.get(name, 0)).font = CELL_FONT

    for c in range(1, len(headers) + 1):
        letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[letter].width = 18


def save_excel(results: list, output_path: Path):
    wb = Workbook()
    wb.remove(wb.active)

    for result in results:
        # Excel sheet name 最长 31 字符
        ws = wb.create_sheet(title=result.label[:31])
        _write_result_sheet(ws, result)

    ws_m = wb.create_sheet(title="metrics")
    _write_metrics_sheet(ws_m, results)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("已保存: %s", output_path)


# ── Provider 构建 ─────────────────────────────────────────────────────────

def make_search(name: str):
    if name == "baidu":
        if not settings.baidu_api_key:
            raise ValueError("BAIDU_API_KEY 未配置")
        return BaiduQianfanProvider(api_key=settings.baidu_api_key)
    if name == "volcengine":
        if not settings.volcengine_api_key:
            raise ValueError("VOLCENGINE_API_KEY 未配置")
        return VolcengineProvider(api_key=settings.volcengine_api_key)
    if name == "metaso":
        if not settings.metaso_api_key:
            raise ValueError("METASO_API_KEY 未配置")
        return MetasoProvider(api_key=settings.metaso_api_key)
    raise ValueError(f"未知搜索源: {name}")


def make_llm(model_key: str):
    cfg = LLM_CONFIGS[model_key]
    key = settings.claude_proxy_api_key
    url = settings.claude_proxy_base_url
    if not key or not url:
        raise ValueError("CLAUDE_PROXY_API_KEY / CLAUDE_PROXY_BASE_URL 未配置")
    return AnthropicCompatProvider(base_url=url, api_key=key, model=cfg["model"]), cfg["label"]


# ── Phase 逻辑 ────────────────────────────────────────────────────────────

PHASE2_COMBOS = [
    ("百度",           ["baidu"]),
    ("火山",           ["volcengine"]),
    ("秘塔",           ["metaso"]),
    ("百度+火山",      ["baidu", "volcengine"]),
    ("百度+秘塔",      ["baidu", "metaso"]),
    ("火山+秘塔",      ["volcengine", "metaso"]),
    ("百度+火山+秘塔", ["baidu", "volcengine", "metaso"]),
]


async def run_phase1():
    print("\n" + "=" * 60)
    print("Phase 1: 百度千帆 × Sonnet vs Haiku")
    print("=" * 60)

    results = []
    for model_key in ["sonnet", "haiku"]:
        base_llm, label = make_llm(model_key)
        llm = CountingLLM(base_llm, label)
        search = CountingSearch(make_search("baidu"))

        result = await run_combination(
            label=f"百度+{label}",
            search_with_limits=[(search, 12)],
            llm=llm,
        )
        results.append(result)
        await search.close()
        await llm.close()

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out = ROOT / "experiments" / "results" / f"phase1_{ts}.xlsx"
    save_excel(results, out)
    _print_summary(results)


async def run_phase2(llm_key: str):
    print("\n" + "=" * 60)
    print(f"Phase 2: 7 种搜索组合 × {LLM_CONFIGS[llm_key]['label']}")
    print("=" * 60)

    results = []
    for combo_label, search_names in PHASE2_COMBOS:
        n = SEARCH_LIMITS[len(search_names)]
        searches = [(CountingSearch(make_search(s)), n) for s in search_names]
        base_llm, llm_label = make_llm(llm_key)
        llm = CountingLLM(base_llm, llm_label)

        result = await run_combination(
            label=combo_label,
            search_with_limits=searches,
            llm=llm,
        )
        results.append(result)

        for s, _ in searches:
            await s.close()
        await llm.close()

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out = ROOT / "experiments" / "results" / f"phase2_{ts}.xlsx"
    save_excel(results, out)
    _print_summary(results)


def _print_summary(results: list):
    print("\n" + "-" * 60)
    print(f"{'组合':<20} {'耗时(s)':>8} {'LLM调用':>8} {'Input':>10} {'Output':>10}")
    print("-" * 60)
    for r in results:
        print(f"{r.label:<20} {r.total_time_s:>8.1f} {r.llm_calls:>8} "
              f"{r.llm_input_tokens:>10} {r.llm_output_tokens:>10}")
        for name, count in r.search_calls.items():
            print(f"  搜索({name}): {count} 次")
    print("-" * 60)


def main():
    parser = argparse.ArgumentParser(description="跨境工具 A/B Benchmark")
    parser.add_argument("--phase", type=int, required=True, choices=[1, 2])
    parser.add_argument("--llm", default="haiku", choices=["sonnet", "haiku"],
                        help="Phase 2 使用的 LLM（默认 haiku）")
    args = parser.parse_args()

    if args.phase == 1:
        asyncio.run(run_phase1())
    else:
        asyncio.run(run_phase2(args.llm))


if __name__ == "__main__":
    main()
